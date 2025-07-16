import openpyxl
from openpyxl import load_workbook
from datetime import datetime
import os
import time

def makeExcel(aantal_sensors):
    # Controleer of het bestand al bestaat
    file_name = "../my_excel_file8.xlsx"
    if os.path.exists(file_name):
        workbook = load_workbook(file_name)
        sheet = workbook["Data"]
        dash = workbook["Dashboard"]

    else:
        # Maak een nieuwe Excel-workbook
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Data"
        dash = workbook.create_sheet()
        dash.title = "Dashboard"

        # Voeg headers toe
        headersDatasheet = ["Datum", "DiverID", "Dichten", "Programmeren", "Graveren", "Controleren", "Inpakken", "Totaal", "Afkeur"]
        headersDashboard = ["Gemiddelde dichttijd", "Gemiddelde programmeertijd", "Gemiddelde graveertijd", "Gemiddelde controleertijd", "Gemiddelde inpaktijd", "Gemiddelde totale tijd", "Afkeurpercentage"]
        sheet.append(headersDatasheet)
        dash.append(headersDashboard)

    DiverID = list(range(1, aantal_sensors + 1))

    # Data-instellingen
    Date = datetime.today().strftime('%d-%m-%Y')
    Tijd_Programmeren = 30

    # Simuleer tijden voor elk station
    for diver_id in DiverID:
        start_time_dichten = time.time()
        time.sleep(2)  # Simuleer verwerkingstijd voor dichten
        Tijd_Dichten = time.time() - start_time_dichten

        start_time_graveren = time.time()
        time.sleep(1)  # Simuleer verwerkingstijd voor graveren
        Tijd_Graveren = time.time() - start_time_graveren

        # Voeg een nieuwe rij toe aan de Excel-sheet
        row = [Date, diver_id, Tijd_Dichten, Tijd_Programmeren, Tijd_Graveren]
        sheet.append(row)

    dash["A2"] = "=AVERAGE(Data!C2:C100)"

    # Sla het bestand op
    workbook.save(file_name)
    print(f"Data succesvol toegevoegd aan {file_name}!")
